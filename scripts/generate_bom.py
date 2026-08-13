#!/usr/bin/env python3
"""Generate MVP_BOM_v1.xlsx from hardware TZ."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "MVP_BOM_v1.xlsx"

HEADERS = [
    "Категория",
    "ID",
    "Наименование",
    "Спецификация / требование",
    "Модуль",
    "Кол-во",
    "Ед.",
    "Приоритет",
    "Пример / артикул",
    "Поставщик (РФ)",
    "Цена ед., USD",
    "Сумма, USD",
    "Срок поставки",
    "Примечание",
]

# (category, id, name, spec, module, qty, unit, priority, example, supplier, price_usd, lead_time, note)
ROWS = [
    # E-01 Compute — Master
    ("E-01 Compute", "E01-001", "Edge SBC", "Raspberry Pi 4 Model B, 4 GB RAM", "MS-CHILL-01", 1, "шт", "Must", "Raspberry Pi 4 4GB", "Амперка / CompTek / DNS", 75, "1–2 нед", "Master edge-host"),
    ("E-01 Compute", "E01-002", "MicroSD / SSD", "SSD ≥32 GB или industrial microSD 32 GB A2", "MS-CHILL-01", 1, "шт", "Must", "Kingston A400 120GB / Samsung PRO Endurance", "DNS / Citilink", 25, "1 нед", ""),
    ("E-01 Compute", "E01-003", "UPS line-interactive", "≥600 VA, USB monitoring, 15 min @ logic load", "MS-CHILL-01", 1, "шт", "Must", "APC Back-UPS 650 / Powercom", "DNS / Citilink", 120, "1–2 нед", "L1 logic + locks"),
    ("E-01 Compute", "E01-004", "Ethernet switch", "5-port Gigabit (lab bench)", "INSTALL", 1, "шт", "Should", "TP-Link TL-SG105", "DNS", 20, "1 нед", "Lab network"),
    ("E-01 Compute", "E01-005", "QR scanner USB", "2D QR для старта сессии", "MS-CHILL-01", 1, "шт", "Should", "Honeywell/YJ4600 class", "AliExpress / Tinko", 40, "2–4 нед", "TBD-04: или pedestal"),
    ("E-01 Compute", "E01-006", "Wi-Fi dongle", "USB AC dual-band fallback", "MS-CHILL-01", 1, "шт", "Optional", "TP-Link Archer T3U", "DNS", 15, "1 нед", ""),

    # E-02 MCU Slaves
    ("E-02 MCU", "E02-001", "Slave MCU board", "ESP32-WROOM-32 или STM32F4 dev board", "ALL SLAVES", 3, "шт", "Must", "ESP32-DevKitC / Nucleo-F446", "AliExpress / ChipDip", 12, "2–3 нед", "1× per module"),
    ("E-02 MCU", "E02-002", "Logic level shifter", "3.3 V ↔ 5 V для RS-485", "ALL SLAVES", 3, "шт", "Must", "TXS0108E module", "ChipDip", 3, "1–2 нед", ""),
    ("E-02 MCU", "E02-003", "Prototype PCB", "Perfboard / custom PCB v0 slave I/O", "ALL SLAVES", 3, "шт", "Should", "Custom fab OR perfboard", "JLCPCB / local", 15, "2–4 нед", ""),

    # E-03 Load cells
    ("E-03 Load cells", "E03-001", "Load cell 3 kg", "Single-point, C3, для слотов produce/meal", "MS-PRODUCE-01", 5, "шт", "Must", "CZL602X 3kg / similar", "AliExpress / Тензоприбор", 8, "3–4 нед", "P-1…P-5"),
    ("E-03 Load cells", "E03-002", "Load cell 15 kg", "Single-point для полки chill", "MS-CHILL-01", 3, "шт", "Must", "CZL604 15kg / HBM SP4", "AliExpress / Тензоприбор", 18, "3–4 нед", "3 полки → C-1…C-5"),
    ("E-03 Load cells", "E03-003", "Load cell 1 kg", "Для meal storage slots", "MS-MEAL-01", 3, "шт", "Must", "CZL602 1kg", "AliExpress", 7, "3–4 нед", "M-1…M-3"),
    ("E-03 Load cells", "E03-004", "HX711 ADC module", "24-bit ADC для тензоячеек", "ALL", 11, "шт", "Must", "HX711 breakout", "AliExpress / ChipDip", 2, "2 нед", "5+3+3 cells"),
    ("E-03 Load cells", "E03-005", "Calibration weights", "Набор 100 g, 500 g, 1 kg F1 class lab", "INSTALL", 1, "компл", "Must", "Гиря 1kg M1", "Metrologiya shop", 40, "2–4 нед", "Калибровка всех слотов"),
    ("E-03 Load cells", "E03-006", "Load cell spare", "10% spare cells", "SPARE", 2, "шт", "Should", "Mixed", "—", 15, "—", "Rework buffer"),

    # E-04 Locks
    ("E-04 Locks", "E04-001", "Electromagnetic lock", "12 V, holding ≥300 N, fail-safe/secure per design", "MS-PRODUCE-01", 1, "шт", "Must", "280 kg holding 12V EM lock", "AliExpress / security shop", 18, "2–3 нед", ""),
    ("E-04 Locks", "E04-002", "Electromagnetic lock", "12 V, для холодильной двери", "MS-CHILL-01", 1, "шт", "Must", "Low heat EM lock 12V", "AliExpress", 22, "2–3 нед", "Low thermal leak"),
    ("E-04 Locks", "E04-003", "Electromagnetic lock", "12 V, storage door meal", "MS-MEAL-01", 1, "шт", "Must", "280 kg holding 12V", "AliExpress", 18, "2–3 нед", ""),
    ("E-04 Locks", "E04-004", "Lock driver MOSFET", "High-side switch + flyback diode board", "ALL", 3, "шт", "Must", "IRF540 + driver", "ChipDip", 5, "1 нед", ""),

    # E-05 Refrigeration
    ("E-05 Refrigeration", "E05-001", "Compressor unit", "Light commercial, R600a, 300–500 L equiv.", "MS-CHILL-01", 1, "компл", "Must", "Danfoss/Embraco class 1/3 HP", "Holodilshik / Ali", 350, "4–8 нед", "L2 dedicated line"),
    ("E-05 Refrigeration", "E05-002", "Evaporator + condenser", "В комплекте или matched pair", "MS-CHILL-01", 1, "компл", "Must", "Matched to E05-001", "Holod supplier", 0, "—", "Included in kit"),
    ("E-05 Refrigeration", "E05-003", "Capillary / thermostat", "Терморегуляция холодильника", "MS-CHILL-01", 1, "компл", "Must", "Capillary thermostat WDF", "Holod supplier", 25, "2 нед", "Or electronic controller"),
    ("E-05 Refrigeration", "E05-004", "Peltier module TEC", "12 V, для meal storage v0 (TBD-03)", "MS-MEAL-01", 2, "шт", "Should", "12706 + heatsink kit", "AliExpress", 30, "2–3 нед", "Alt: mini compressor"),
    ("E-05 Refrigeration", "E05-005", "PU insulation panel", "40 mm, для камеры chill", "MS-CHILL-01", 4, "лист", "Must", "600×800×40 PU", "Local fab", 40, "2 нед", "Cut to size"),
    ("E-05 Refrigeration", "E05-006", "Door gasket", "Magnetic profile per door", "MS-CHILL-01", 1, "м", "Must", "PVC magnetic gasket 12×4", "Holod supplier", 8, "1 нед", ""),

    # E-06 Induction
    ("E-06 Induction", "E06-001", "Induction module", "Single zone 800–1500 W, serial/PWM control", "MS-MEAL-01", 1, "шт", "Must", "Commercial induction module 1kW", "AliExpress / kitchen equip", 120, "3–5 нед", "L3 16A line"),
    ("E-06 Induction", "E06-002", "E-stop mushroom", "NC, 22 mm, cutout", "MS-MEAL-01", 1, "шт", "Must", "XB2-BS542", "ChipDip / IEK", 12, "1 нед", "Hard interrupt induction"),
    ("E-06 Induction", "E06-003", "Induction-compatible container", "PP bowl Ø180 mm or adapter plate", "MS-MEAL-01", 10, "шт", "Must", "PP meal container + induction plate", "Packaging supplier", 2, "1 нед", "Lab + spare"),
    ("E-06 Induction", "E06-004", "Exhaust fan", "120 mm, 12 V, post-heat cooling", "MS-MEAL-01", 1, "шт", "Must", "120mm DC fan", "DNS", 8, "1 нед", ""),

    # E-07 Bus
    ("E-07 Bus", "E07-001", "RS-485 transceiver", "MAX485 / SN65HVD1782", "MS-CHILL-01", 1, "шт", "Must", "MAX485 module", "ChipDip", 3, "1 нед", "Bus master (TBD-02)"),
    ("E-07 Bus", "E07-002", "RS-485 transceiver", "Slave transceivers", "SLAVES", 2, "шт", "Must", "MAX485 module", "ChipDip", 3, "1 нед", ""),
    ("E-07 Bus", "E07-003", "Twisted pair cable", "Shielded 2×0.34 mm² + drain", "INSTALL", 10, "м", "Must", "КВББШв 2×0.5 or RS-485 cable", "Electro shop", 15, "1 нед", "Inter-module ≤3 m"),
    ("E-07 Bus", "E07-004", "Termination resistor", "120 Ω 1/4 W", "INSTALL", 2, "шт", "Must", "120Ω", "ChipDip", 1, "1 нед", "Bus ends"),
    ("E-07 Bus", "E07-005", "CAN transceiver alt.", "TJA1050 — if CAN chosen TBD-02", "ALL", 3, "шт", "Optional", "TJA1050 module", "ChipDip", 4, "1 нед", "Alternative to RS-485"),

    # E-08 Mechanical
    ("E-08 Mechanical", "E08-001", "Aluminum profile", "40×40 mm, каркас модулей", "ALL", 60, "м", "Must", "2020/4040 profile", "Local metal shop", 180, "2–3 нед", "~20 m per module"),
    ("E-08 Mechanical", "E08-002", "Steel sheet", "1.5 mm, обшивка", "ALL", 8, "лист", "Must", "1250×2500×1.5", "Metal base", 120, "2 нед", ""),
    ("E-08 Mechanical", "E08-003", "Produce bunker", "AISI 304, angle ~45°, welded", "MS-PRODUCE-01", 2, "шт", "Must", "Custom fab", "Local stainless shop", 150, "3–4 нед", "P-1, P-2"),
    ("E-08 Mechanical", "E08-004", "Removable tray", "Food-grade PP/304, produce slots", "MS-PRODUCE-01", 3, "шт", "Must", "Custom", "Local fab", 30, "2 нед", "P-3…P-5"),
    ("E-08 Mechanical", "E08-005", "Chill shelves", "Perforated SS, adjustable", "MS-CHILL-01", 4, "шт", "Must", "600×400 shelf", "Local fab", 40, "2 нед", ""),
    ("E-08 Mechanical", "E08-006", "Door hinges + closer", "110° open, soft close", "ALL", 3, "компл", "Must", "Overlay hinge set", "Furniture hardware", 45, "1 нед", "1 set per module"),
    ("E-08 Mechanical", "E08-007", "Transparent door insert", "Polycarbonate window optional", "MS-PRODUCE-01", 1, "шт", "Optional", "PC sheet 4mm", "Plastic shop", 25, "1 нед", ""),
    ("E-08 Mechanical", "E08-008", "Base frame", "Optional common plinth", "INSTALL", 1, "компл", "Optional", "Steel tube frame", "Local fab", 80, "2 нед", ""),
    ("E-08 Mechanical", "E08-009", "Slot dividers", "Marked zones C-1…C-10", "MS-CHILL-01", 10, "шт", "Should", "PP dividers", "Local", 20, "1 нед", ""),
    ("E-08 Mechanical", "E08-010", "Cable channel", "40×25 mm between modules", "INSTALL", 3, "м", "Must", "PVC channel", "Electro shop", 12, "1 нед", ""),

    # E-09 Sensors
    ("E-09 Sensors", "E09-001", "NTC temperature sensor", "10 kΩ, waterproof probe", "MS-CHILL-01", 2, "шт", "Must", "DS18B20 waterproof / NTC", "ChipDip", 4, "1 нед", "Cabinet + evaporator"),
    ("E-09 Sensors", "E09-002", "NTC / IR temp", "Plate temp induction zone", "MS-MEAL-01", 1, "шт", "Must", "MLX90614 or NTC plate", "AliExpress", 12, "2–3 нед", ""),
    ("E-09 Sensors", "E09-003", "Door reed switch", "NC/NO magnetic", "ALL", 4, "шт", "Must", "MC-38", "ChipDip", 2, "1 нед", ""),
    ("E-09 Sensors", "E09-004", "Ambient temp/humidity", "Optional produce cabinet", "MS-PRODUCE-01", 1, "шт", "Optional", "SHT31 module", "ChipDip", 6, "1 нед", ""),
    ("E-09 Sensors", "E09-005", "Current sensor", "Induction pan detect", "MS-MEAL-01", 1, "шт", "Should", "ACS712 30A", "ChipDip", 4, "1 нед", ""),
    ("E-09 Sensors", "E09-006", "USB camera", "720p optional produce log TBD-07", "MS-PRODUCE-01", 1, "шт", "Optional", "Logitech C270 class", "DNS", 25, "1 нед", ""),

    # E-10 PSU
    ("E-10 PSU", "E10-001", "24 V PSU central", "10 A, mean well class", "MS-CHILL-01", 1, "шт", "Must", "Mean Well LRS-150-24", "ChipDip / electro", 35, "1–2 нед", "Logic bus + locks"),
    ("E-10 PSU", "E10-002", "12 V DC sub-PSU", "5 A for locks/ fans", "MS-CHILL-01", 1, "шт", "Must", "LRS-75-12", "ChipDip", 25, "1–2 нед", ""),
    ("E-10 PSU", "E10-003", "5 V DC buck", "3 A for MCU/edge peripherals", "ALL", 3, "шт", "Must", "LM2596 module", "ChipDip", 3, "1 нед", ""),
    ("E-10 PSU", "E10-004", "Surge protector PDU", "Lab rack protection", "INSTALL", 1, "шт", "Should", "PDU 6 outlet", "DNS", 30, "1 нед", ""),

    # E-11 Cables & connectors
    ("E-11 Cables", "E11-001", "Power cable 3×1.5", "L1/L2/L3 to modules", "INSTALL", 15, "м", "Must", "ВВГнг 3×1.5", "Electro shop", 25, "1 нед", ""),
    ("E-11 Cables", "E11-002", "WAGO connectors", "221 series assortment", "INSTALL", 1, "компл", "Must", "WAGO 221 kit", "Electro shop", 35, "1 нед", ""),
    ("E-11 Cables", "E11-003", "XT30/XT60 connectors", "Modular power service", "ALL", 6, "пар", "Should", "XT30", "AliExpress", 10, "2 нед", ""),
    ("E-11 Cables", "E11-004", "Ethernet CAT5e", "Edge to router", "INSTALL", 5, "м", "Must", "Patch + cable", "DNS", 8, "1 нед", ""),
    ("E-11 Cables", "E11-005", "Dupont / JST wire kit", "Prototyping", "ALL", 1, "компл", "Must", "Assorted kit", "AliExpress", 15, "2 нед", ""),

    # E-12 Food contact & consumables
    ("E-12 Food", "E12-001", "PP food containers meal", "350 ml with lid", "MS-MEAL-01", 50, "шт", "Must", "Meal prep container", "Pack wholesale", 25, "1 нед", "Lab testing"),
    ("E-12 Food", "E12-002", "Produce mesh bags", "1 kg / 500 g for SKU demo", "MS-PRODUCE-01", 20, "шт", "Must", "Net bags", "Pack shop", 5, "1 нед", ""),
    ("E-12 Food", "E12-003", "Dairy sample SKU", "Milk/kefir/tvorog for calibration", "MS-CHILL-01", 10, "шт", "Must", "Retail packs", "Store", 15, "1 день", "Calibration only"),
    ("E-12 Food", "E12-004", "Sanitizer / food-safe cleaner", "Lab hygiene", "INSTALL", 1, "компл", "Should", "Food contact cleaner", "Retail", 10, "1 день", ""),

    # INSTALL — electrical protection
    ("INSTALL", "INS-001", "Circuit breaker 10A", "L2 compressor", "INSTALL", 1, "шт", "Must", "ABB/IEK 1P C10", "Electro shop", 8, "1 нед", ""),
    ("INSTALL", "INS-002", "Circuit breaker 16A", "L3 induction", "INSTALL", 1, "шт", "Must", "ABB/IEK 1P C16", "Electro shop", 8, "1 нед", ""),
    ("INSTALL", "INS-003", "Circuit breaker 6A", "L1 logic + UPS input", "INSTALL", 1, "шт", "Must", "ABB/IEK 1P C6", "Electro shop", 8, "1 нед", ""),
    ("INSTALL", "INS-004", "RCD 30mA", "Recommended L2/L3", "INSTALL", 1, "шт", "Should", "IEK RCD 30mA", "Electro shop", 20, "1 нед", ""),
    ("INSTALL", "INS-005", "Label kit", "L1/L2/L3, slot IDs, QR serial", "INSTALL", 1, "компл", "Must", "Industrial labels", "Print shop", 15, "1 нед", ""),

    # SPARE
    ("SPARE", "SPR-001", "Spare EM lock", "Common lock type", "SPARE", 1, "шт", "Should", "Same as E04", "—", 20, "—", ""),
    ("SPARE", "SPR-002", "Spare MCU board", "ESP32 devkit", "SPARE", 1, "шт", "Should", "ESP32-DevKitC", "—", 12, "—", ""),
    ("SPARE", "SPR-003", "Misc fasteners kit", "M5/M6 stainless assortment", "ALL", 1, "компл", "Must", "Assortment box", "Fix shop", 25, "1 нед", ""),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4D4D")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CAT_FILL = PatternFill("solid", fgColor="E8F0F0")
MUST_FILL = PatternFill("solid", fgColor="FFFFFF")
OPT_FILL = PatternFill("solid", fgColor="FFF8E8")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build():
    wb = Workbook()

    # --- Sheet 1: BOM ---
    ws = wb.active
    ws.title = "BOM"

    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = 2
    total = 0.0
    must_total = 0.0
    prev_cat = None

    for r in ROWS:
        cat, rid, name, spec, module, qty, unit, priority, example, supplier, price, lead, note = r
        line_total = qty * price
        total += line_total
        if priority == "Must":
            must_total += line_total

        ws.append([
            cat, rid, name, spec, module, qty, unit, priority,
            example, supplier, price, line_total, lead, note,
        ])

        if cat != prev_cat:
            row_fill = CAT_FILL
            prev_cat = cat
        elif priority == "Optional":
            row_fill = OPT_FILL
        else:
            row_fill = MUST_FILL

        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col).fill = row_fill

        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col).border = BORDER
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        # Price columns numeric
        ws.cell(row=row_num, column=11).number_format = "#,##0.00"
        ws.cell(row=row_num, column=12).number_format = "#,##0.00"

        row_num += 1

    # Totals row
    ws.append([])
    tr = row_num + 1
    ws.cell(row=tr, column=10, value="ИТОГО (all lines):").font = Font(bold=True)
    ws.cell(row=tr, column=12, value=total).font = Font(bold=True)
    ws.cell(row=tr, column=12).number_format = "#,##0.00"
    ws.cell(row=tr + 1, column=10, value="ИТОГО (Must only):").font = Font(bold=True)
    ws.cell(row=tr + 1, column=12, value=must_total).font = Font(bold=True)
    ws.cell(row=tr + 1, column=12).number_format = "#,##0.00"

    widths = [14, 10, 28, 42, 16, 8, 6, 10, 28, 22, 12, 12, 12, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary by category ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Категория", "Позиций", "Сумма USD", "Must USD"])
    for col in range(1, 5):
        ws2.cell(row=1, column=col).fill = HEADER_FILL
        ws2.cell(row=1, column=col).font = HEADER_FONT

    from collections import defaultdict
    cat_stats = defaultdict(lambda: {"count": 0, "sum": 0.0, "must": 0.0})
    for r in ROWS:
        cat = r[0]
        qty, price, priority = r[5], r[10], r[7]
        line = qty * price
        cat_stats[cat]["count"] += 1
        cat_stats[cat]["sum"] += line
        if priority == "Must":
            cat_stats[cat]["must"] += line

    sr = 2
    for cat in sorted(cat_stats.keys()):
        s = cat_stats[cat]
        ws2.append([cat, s["count"], s["sum"], s["must"]])
        ws2.cell(row=sr, column=3).number_format = "#,##0.00"
        ws2.cell(row=sr, column=4).number_format = "#,##0.00"
        sr += 1

    ws2.append([])
    ws2.cell(row=sr + 1, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=sr + 1, column=2, value=sum(x["count"] for x in cat_stats.values()))
    ws2.cell(row=sr + 1, column=3, value=total).font = Font(bold=True)
    ws2.cell(row=sr + 1, column=4, value=must_total).font = Font(bold=True)

    for i, w in enumerate([18, 10, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 3: By module ---
    ws3 = wb.create_sheet("By Module")
    ws3.append(["Модуль", "Позиций", "Сумма USD"])
    for col in range(1, 4):
        ws3.cell(row=1, column=col).fill = HEADER_FILL
        ws3.cell(row=1, column=col).font = HEADER_FONT

    mod_stats = defaultdict(lambda: {"count": 0, "sum": 0.0})
    for r in ROWS:
        mod = r[4]
        line = r[5] * r[10]
        mod_stats[mod]["count"] += 1
        mod_stats[mod]["sum"] += line

    mr = 2
    for mod in sorted(mod_stats.keys()):
        s = mod_stats[mod]
        ws3.append([mod, s["count"], s["sum"]])
        ws3.cell(row=mr, column=3).number_format = "#,##0.00"
        mr += 1

    for i, w in enumerate([20, 10, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 4: Meta ---
    ws4 = wb.create_sheet("Meta")
    meta = [
        ("Документ", "MVP_BOM_v1.xlsx"),
        ("Версия", "1.0"),
        ("Дата", "2026-08-13"),
        ("Проект", "Micro Shop MVP"),
        ("Рынок", "РФ"),
        ("Базовое ТЗ", "MVP_TZ_HARDWARE_v1.md"),
        ("Валюта оценки", "USD (ориентир закупки)"),
        ("", ""),
        ("Примечания", ""),
        ("1", "Цены — ориентировочные для budgeting v0, не коммерческие оферты"),
        ("2", "TBD-02: RS-485 (default) — CAN позиции помечены Optional"),
        ("3", "TBD-03: Meal cold — Peltier (Should); alt mini compressor +$200"),
        ("4", "Mechanical fab (E-08) — сильно зависит от локального производства"),
        ("5", "Must total — минимальный budget для старта закупки"),
        ("6", "Не включено: работа инженеров, доставка, НДС, lab аренда"),
    ]
    for row in meta:
        ws4.append(list(row))
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 60

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Total: ${total:,.2f} | Must: ${must_total:,.2f}")


if __name__ == "__main__":
    build()
